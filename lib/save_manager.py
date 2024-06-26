import os
import json
from openpyxl import Workbook
from jinja2 import Environment, FileSystemLoader
from config.settings import ResultSetting, DatabaseSetting
from model.result import Result, QueryResult
from utils.utils import Hashing
from utils.ansiprint import AnsiPrint
from middle import __VERSION__

class SaveManager:
    def __init__(self) -> None:
        self._results = ResultSetting() #Settings.setting["Results"]
        self._output = self._results.output
        
        if self._output is None:
            home = os.path.expanduser("~")
            self._output = os.path.join(home,'.local','share','mapXplore')
            self._results.output = self._output

    def _create_directory(self, directory:str):
        if not os.path.exists(directory):
            os.makedirs(directory)

    def _write_content(self, content, name):
        
        if isinstance(content, bytes):
            with(open(name,'wb')) as fb:
                fb.write(content)
        else:
            with(open(name,'w')) as fb:
                fb.write(content)

    def save(self, content, format, name=None, directory:str=None) -> str:
        output_directory = directory if directory is not None else self._output
        self._create_directory(output_directory)
        if name is None:
            name = Hashing.get_md5(content)
            if format is not None:
                name+=f".{format}"
        
        path = os.path.join(output_directory, f"{name}.{format}")

        self._write_content(content, path)
        return path
    
    def convert_content_to_plain(self, items:QueryResult)->str:
        self._output_dir = ResultSetting().get_folder_output(ResultSetting().format)
        fileType = ResultSetting().get_FileType
        if fileType == ResultSetting.ExportFileType.Excel:
            return self.convert_to_csv(items)
        elif fileType == ResultSetting.ExportFileType.JSON:
            return self.convert_to_json(items)
        else:
            files = []
            for item in items:
                html = self.convert_to_html(item.results)
                file = self.save(html, 'html', directory=self._output_dir)
                files.append({"name":item.value, "file":os.path.basename(file)})
            return self._create_index(files)

    def _create_index(self, files:list[str]) ->str:
        env = Environment(loader=FileSystemLoader("."))
        index_template = env.get_template('/templates/index.html')
        html = index_template.render(files=files)
        output = self.save(html, 'html', 'index', self._output_dir)
        return output
    
    def convert_to_html(self, items:Result):
        env = Environment(loader=FileSystemLoader("."))
        template = env.get_template('/templates/result.html')
        output = template.render(headers=items.headers, 
                                rows=items.rows, 
                                formatted_rows=items.formatted_rows,
                                table_name=items.table_name)
        return output
    
    def convert_to_json(self, items:Result):
        self._create_directory(self._output_dir)
        data = []
        file_name = os.path.join(self._output_dir, Hashing.get_md5(''.join([str(i.criterial) for i in items])))+".json"
        for item in items:
            information = {item.value:[]}
            
            headers = item.results.headers
            for row in item.results.rows:
                for index, value in enumerate(row):
                    information[item.value].append({headers[index]:value})
            
            data.append(information)
        
        with open(file_name, "w") as outfile:
            outfile.write(json.dumps(data, indent=4))
        
        return file_name

    def convert_to_csv(self, items:QueryResult) -> str:
        try:
            self._create_directory(self._output_dir)
            file_name = os.path.join(self._output_dir, Hashing.get_md5(''.join([str(i.criterial) for i in items])))+".xlsx"
            wb = Workbook()
            for index, result in enumerate(items):
                ws = None
                if index == 0:
                    ws = wb.active
                    ws.title = result.value
                else:
                    ws = wb.create_sheet(result.value)

                for col, h in enumerate(result.results.headers, start=1):
                    ws.cell(row=1, column=col, value=h)
                
                for row_idx,row_data in enumerate(result.results.rows, start=2):
                    for col_idx, value in enumerate(row_data, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
            
            wb.save(file_name)
            return file_name
        except Exception as e:
            AnsiPrint.print_debug(e)
    
    def save_content(self, items:list[Result]) -> str:
        explorer_folder = ResultSetting().get_folder_output('html')
        env = Environment(loader=FileSystemLoader("."))
        template = env.get_template('/templates/explorer.html')
        output = template.render(options=items, version=__VERSION__)
        file = self.save(output, 'html', directory=explorer_folder, name='explorer')
        return file
        